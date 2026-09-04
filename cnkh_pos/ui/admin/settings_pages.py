from __future__ import annotations

import json
import os
from datetime import date, datetime
from pathlib import Path

from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QPixmap
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from cnkh_pos.config import RECEIPT_QR_IMAGE_NAME, AppPaths
from cnkh_pos.database.connection import Database
from cnkh_pos.database.migrations import utc_now_text
from cnkh_pos.services.auth import AuthenticatedUser
from cnkh_pos.services.catalog import CategoryService
from cnkh_pos.services.daily_closing import DailyClosingService
from cnkh_pos.services.document_numbers import (
    DEFAULT_DOCUMENT_PREFIXES,
    document_prefixes,
    save_document_prefixes,
)
from cnkh_pos.services.lan_sync_server import (
    DEFAULT_PORT,
    detect_lan_ip,
    get_active_server,
    start_global_server,
    stop_global_server,
)
from cnkh_pos.services.money import format_myr, rm_to_cents
from cnkh_pos.services.printing import (
    WINDOWS_DEFAULT_PRINTER,
    resolve_printer_target,
)
from cnkh_pos.services.reports import ReportService
from cnkh_pos.ui.dialogs.sync_pair_dialog import open_sync_pair_dialog


class CategoryDialog(QDialog):
    def __init__(self, database: Database, user: AuthenticatedUser, parent=None):
        super().__init__(parent)
        self.database = database
        self.user = user
        self.setWindowTitle("Category Management / 分类管理")
        self.setMinimumSize(680, 500)
        root = QVBoxLayout(self)
        title = QLabel("分类管理")
        title.setObjectName("PageTitle")
        root.addWidget(title)
        body = QHBoxLayout()
        self.list = QListWidget()
        body.addWidget(self.list, 2)
        detail = QVBoxLayout()
        detail.addWidget(QLabel("Category Name"))
        self.name = QLineEdit()
        detail.addWidget(self.name)
        self.count = QLabel("Product Count: 0")
        detail.addWidget(self.count)
        detail.addStretch(1)
        body.addLayout(detail, 3)
        root.addLayout(body, 1)
        actions = QHBoxLayout()
        for text, callback, style in (
            ("新增分类", self.add, "PrimaryButton"),
            ("修改分类", self.rename, "WarningButton"),
            ("删除分类", self.delete, "DangerButton"),
            ("关闭", self.accept, ""),
        ):
            button = QPushButton(text)
            button.setObjectName(style)
            button.clicked.connect(callback)
            actions.addWidget(button)
        root.addLayout(actions)
        self.list.currentItemChanged.connect(self._selected)
        self.refresh()

    def refresh(self) -> None:
        self.list.clear()
        conn = self.database.connect(readonly=True)
        try:
            rows = conn.execute(
                """SELECT c.id,c.name,COUNT(p.id) count FROM categories c
                   LEFT JOIN products p ON p.category_id=c.id AND p.is_deleted=0
                   GROUP BY c.id ORDER BY c.name"""
            ).fetchall()
            for row in rows:
                from PySide6.QtWidgets import QListWidgetItem

                item = QListWidgetItem(str(row["name"]))
                item.setData(
                    Qt.ItemDataRole.UserRole, (int(row["id"]), int(row["count"]))
                )
                self.list.addItem(item)
        finally:
            conn.close()

    def _selected(self, current, previous) -> None:
        del previous
        if current:
            category_id, count = current.data(Qt.ItemDataRole.UserRole)
            self.name.setText(current.text())
            self.count.setText(f"Product Count: {count}")

    def add(self) -> None:
        name = self.name.text().strip()
        if name:
            CategoryService(self.database).add(name, admin_id=self.user.id)
            self.refresh()

    def rename(self) -> None:
        item = self.list.currentItem()
        if item and self.name.text().strip():
            CategoryService(self.database).rename(
                item.data(Qt.ItemDataRole.UserRole)[0],
                self.name.text(),
                admin_id=self.user.id,
            )
            self.refresh()

    def delete(self) -> None:
        item = self.list.currentItem()
        if not item:
            return
        category_id, count = item.data(Qt.ItemDataRole.UserRole)
        if (
            QMessageBox.question(
                self, "Delete Category", f"删除后 {count} 个商品会转为未分类。继续？"
            )
            == QMessageBox.StandardButton.Yes
        ):
            CategoryService(self.database).delete(category_id, admin_id=self.user.id)
            self.refresh()


class QuickAmountsWidget(QWidget):
    def __init__(self, database: Database):
        super().__init__()
        self.database = database
        root = QVBoxLayout(self)
        heading = QHBoxLayout()
        title = QLabel("金额快捷按钮设置")
        title.setObjectName("SectionTitle")
        heading.addWidget(title)
        heading.addStretch(1)
        for text, callback, style in (
            ("＋ 新增", self.add, "PrimaryButton"),
            ("修改", self.edit, "WarningButton"),
            ("启用/禁用", self.toggle, ""),
            ("↑", lambda: self.move(-1), ""),
            ("↓", lambda: self.move(1), ""),
            ("删除", self.delete, "DangerButton"),
        ):
            button = QPushButton(text)
            button.setObjectName(style)
            button.clicked.connect(callback)
            heading.addWidget(button)
        root.addLayout(heading)
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["#", "按钮金额 (RM)", "状态", "顺序"])
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.table.verticalHeader().setVisible(False)
        root.addWidget(self.table)
        self.refresh()

    def refresh(self) -> None:
        conn = self.database.connect(readonly=True)
        try:
            rows = conn.execute(
                "SELECT id,amount_cents,is_enabled,sort_order FROM quick_amounts ORDER BY sort_order,id"
            ).fetchall()
        finally:
            conn.close()
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, value in enumerate(
                (
                    row["id"],
                    f"{row['amount_cents'] / 100:.2f}",
                    "启用" if row["is_enabled"] else "禁用",
                    row["sort_order"],
                )
            ):
                self.table.setItem(r, c, QTableWidgetItem(str(value)))

    def selected_id(self) -> int | None:
        row = self.table.currentRow()
        return None if row < 0 else int(self.table.item(row, 0).text())

    def add(self) -> None:
        amount, ok = QInputDialog.getDouble(
            self, "Quick Amount", "Amount RM", 10, 0.01, 100000, 2
        )
        if ok:
            with self.database.transaction() as conn:
                order = int(
                    conn.execute(
                        "SELECT COALESCE(MAX(sort_order),0)+10 FROM quick_amounts"
                    ).fetchone()[0]
                )
                conn.execute(
                    "INSERT INTO quick_amounts(amount_cents,is_enabled,sort_order) VALUES (?,1,?)",
                    (rm_to_cents(amount), order),
                )
            self.refresh()

    def edit(self) -> None:
        item_id = self.selected_id()
        if item_id is None:
            return
        amount, ok = QInputDialog.getDouble(
            self, "Quick Amount", "Amount RM", 10, 0.01, 100000, 2
        )
        if ok:
            with self.database.transaction() as conn:
                conn.execute(
                    "UPDATE quick_amounts SET amount_cents=? WHERE id=?",
                    (rm_to_cents(amount), item_id),
                )
            self.refresh()

    def toggle(self) -> None:
        item_id = self.selected_id()
        if item_id is not None:
            with self.database.transaction() as conn:
                conn.execute(
                    "UPDATE quick_amounts SET is_enabled=1-is_enabled WHERE id=?",
                    (item_id,),
                )
            self.refresh()

    def delete(self) -> None:
        item_id = self.selected_id()
        if item_id is not None:
            with self.database.transaction() as conn:
                conn.execute("DELETE FROM quick_amounts WHERE id=?", (item_id,))
            self.refresh()

    def move(self, direction: int) -> None:
        row = self.table.currentRow()
        if row < 0:
            return
        target_row = row + direction
        if target_row < 0 or target_row >= self.table.rowCount():
            return
        current_id = int(self.table.item(row, 0).text())
        target_id = int(self.table.item(target_row, 0).text())
        current_order = int(self.table.item(row, 3).text())
        target_order = int(self.table.item(target_row, 3).text())
        with self.database.transaction() as conn:
            conn.execute(
                "UPDATE quick_amounts SET sort_order=? WHERE id=?",
                (target_order, current_id),
            )
            conn.execute(
                "UPDATE quick_amounts SET sort_order=? WHERE id=?",
                (current_order, target_id),
            )
        self.refresh()
        self.table.selectRow(target_row)


class ReceiptSettingsWidget(QWidget):
    def __init__(self, database: Database, user: AuthenticatedUser):
        super().__init__()
        self.database = database
        self.user = user
        root = QHBoxLayout(self)
        form = QFormLayout()
        self.store = QLineEdit("CNKH Hardware")
        self.address = QTextEdit()
        self.phone = QLineEdit()
        self.footer = QTextEdit("Thank you / 谢谢光临")
        self.notes = QTextEdit()
        self.printer = QComboBox()
        for label, widget in (
            ("Store Name", self.store),
            ("Address", self.address),
            ("Phone", self.phone),
            ("Footer", self.footer),
            ("Notes", self.notes),
            ("Windows Printer", self.printer),
        ):
            form.addRow(label, widget)
            if hasattr(widget, "textChanged"):
                widget.textChanged.connect(self.update_preview)

        qr_section = QLabel("DuitNow 收款码（仅管理员可改） / DuitNow QR (Admin only)")
        qr_section.setObjectName("SectionTitle")
        qr_section.setWordWrap(True)
        form.addRow(qr_section)
        self.qr_enabled = QCheckBox(
            "在收据打印 DuitNow 收款码 / Print DuitNow QR on receipt"
        )
        self.qr_enabled.setObjectName("ReceiptQrEnabled")
        self.qr_enabled.stateChanged.connect(self.update_preview)
        form.addRow("收据打印 / Print", self.qr_enabled)
        qr_row = QHBoxLayout()
        upload_qr = QPushButton("上传 / 替换 QR")
        upload_qr.setObjectName("ReceiptQrUploadButton")
        upload_qr.setProperty("acceptanceName", "ReceiptQrUploadButton")
        upload_qr.clicked.connect(self.upload_qr)
        clear_qr = QPushButton("清除 QR")
        clear_qr.setObjectName("ReceiptQrClearButton")
        clear_qr.setProperty("acceptanceName", "ReceiptQrClearButton")
        clear_qr.clicked.connect(self.clear_qr)
        qr_row.addWidget(upload_qr)
        qr_row.addWidget(clear_qr)
        form.addRow("管理（仅管理员）", qr_row)
        self.qr_preview = QLabel("No QR image")
        self.qr_preview.setObjectName("ReceiptQrPreview")
        self.qr_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.qr_preview.setFixedSize(120, 120)
        self.qr_preview.setStyleSheet(
            "background:#F7FAFC; border:1px solid #DCE3EC; color:#5B6B7C;"
        )
        self.qr_preview.setScaledContents(False)
        form.addRow("预览 / Preview", self.qr_preview)
        self._qr_source_path: str | None = None

        bt_note = QLabel("手机端可选蓝牙热敏小票机（默认关）；PC 仍用 Windows/USB 打印机。\nMobile optional Bluetooth thermal printer is separate — PC uses Windows/USB printers.")
        bt_note.setWordWrap(True)
        form.addRow("", bt_note)
        save = QPushButton("保存 Receipt Settings")
        save.setObjectName("SuccessButton")
        save.clicked.connect(self.save)
        self.save_button = save
        test = QPushButton("Test Print")
        test.setObjectName("PrimaryButton")
        test.clicked.connect(self.test_print)
        self.test_button = test
        form.addRow(save, test)
        refresh_printers = QPushButton("刷新打印机列表")
        refresh_printers.clicked.connect(self.refresh_printers)
        form.addRow("", refresh_printers)
        root.addLayout(form, 3)
        preview_box = QVBoxLayout()
        preview_box.addWidget(QLabel("80mm Receipt Live Preview"))
        self.preview = QTextEdit()
        self.preview.setReadOnly(True)
        self.preview.setMinimumWidth(310)
        self.preview.setStyleSheet(
            "font-family:Consolas; background:white; border:1px solid #DCE3EC;"
        )
        preview_box.addWidget(self.preview)
        root.addLayout(preview_box, 2)
        self.refresh_printers()
        self.load()
        self.update_preview()

    def refresh_printers(self) -> None:
        from PySide6.QtPrintSupport import QPrinterInfo

        selected = self.printer.currentData() if self.printer.count() else None
        self.printer.clear()
        self.printer.addItem("Select a printer / 请选择打印机", None)
        self.printer.addItem(
            "Windows Default Printer / 系统默认", WINDOWS_DEFAULT_PRINTER
        )
        for name in QPrinterInfo.availablePrinterNames():
            self.printer.addItem(name, name)
        index = self.printer.findData(selected)
        if index < 0 and selected not in (None, WINDOWS_DEFAULT_PRINTER):
            self.printer.addItem(f"Unavailable / 已离线：{selected}", selected)
            index = self.printer.count() - 1
        self.printer.setCurrentIndex(max(0, index))

    def load(self) -> None:
        conn = self.database.connect(readonly=True)
        try:
            row = conn.execute(
                "SELECT value_json FROM settings WHERE key='receipt'"
            ).fetchone()
        finally:
            conn.close()
        if row is None:
            return
        try:
            value = json.loads(row["value_json"])
        except (TypeError, ValueError):
            return
        self.store.setText(str(value.get("store_name", "CNKH Hardware")))
        self.address.setPlainText(str(value.get("address", "")))
        self.phone.setText(str(value.get("phone", "")))
        self.footer.setPlainText(str(value.get("footer", "Thank you / 谢谢光临")))
        self.notes.setPlainText(str(value.get("notes", "")))
        qr_flag = value.get("qr_enabled", False)
        if isinstance(qr_flag, str):
            self.qr_enabled.setChecked(
                qr_flag.strip().lower() in {"1", "true", "yes", "on"}
            )
        else:
            self.qr_enabled.setChecked(bool(qr_flag))
        self._qr_source_path = str(value.get("qr_image", "") or "") or None
        self._refresh_qr_preview()
        mode = str(value.get("printer_mode", "")).upper()
        name = str(value.get("printer_name", "")).strip()
        if mode == "DEFAULT":
            target = WINDOWS_DEFAULT_PRINTER
        elif mode == "NAMED" or (not mode and name):
            target = name
        else:
            target = None
        index = self.printer.findData(target)
        if index < 0 and target not in (None, WINDOWS_DEFAULT_PRINTER):
            self.printer.addItem(f"Unavailable / 已离线：{target}", target)
            index = self.printer.count() - 1
        self.printer.setCurrentIndex(max(0, index))

    def update_preview(self) -> None:
        qr_note = ""
        if self.qr_enabled.isChecked() and self._current_qr_path() is not None:
            qr_note = f"\n{'[QR image attached]':^32}"
        self.preview.setPlainText(
            f"{self.store.text():^32}\n{self.address.toPlainText():^32}\n{self.phone.text():^32}\n"
            f"{'-' * 32}\nReceipt: CNKH20260809-001\nCashier: Admin\n{'-' * 32}\n"
            f"PVC Pipe 20mm      RM 9.00\nHammer 2lb         RM 15.90\n{'-' * 32}\n"
            f"TOTAL              RM 24.90\n{'-' * 32}\n{self.footer.toPlainText():^32}\n{self.notes.toPlainText():^32}"
            f"{qr_note}"
        )

    def save(self) -> None:
        selected = self.printer.currentData()
        if selected is None:
            QMessageBox.warning(
                self,
                "Receipt Settings",
                "请选择 Windows 默认打印机或一台指定打印机。",
            )
            return
        mode = "DEFAULT" if selected == WINDOWS_DEFAULT_PRINTER else "NAMED"
        paths = AppPaths.default()
        paths.ensure_directories()
        qr_image_key = ""
        source = self._qr_source_path
        if source:
            source_path = Path(source)
            if source_path.is_file():
                target = paths.receipt_qr_image
                # Keep a stable Assets/receipt_qr.png (or .jpg) name for printing.
                suffix = source_path.suffix.lower()
                if suffix in {".jpg", ".jpeg"}:
                    target = paths.assets / "receipt_qr.jpg"
                elif suffix == ".png":
                    target = paths.assets / RECEIPT_QR_IMAGE_NAME
                else:
                    target = paths.receipt_qr_image
                if source_path.resolve() != target.resolve():
                    import shutil

                    shutil.copy2(source_path, target)
                qr_image_key = target.name
                self._qr_source_path = str(target)
            elif (paths.assets / Path(source).name).is_file():
                qr_image_key = Path(source).name
            elif source in {RECEIPT_QR_IMAGE_NAME, "receipt_qr.jpg"}:
                qr_image_key = source
        value = {
            "store_name": self.store.text(),
            "address": self.address.toPlainText(),
            "phone": self.phone.text(),
            "footer": self.footer.toPlainText(),
            "notes": self.notes.toPlainText(),
            "printer_mode": mode,
            "printer_name": "" if mode == "DEFAULT" else str(selected),
            "qr_enabled": bool(self.qr_enabled.isChecked()),
            "qr_image": qr_image_key,
        }
        with self.database.transaction() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO settings(key,value_json,updated_at,updated_by) VALUES ('receipt',?,?,?)",
                (json.dumps(value, ensure_ascii=False), utc_now_text(), self.user.id),
            )
        self._refresh_qr_preview()
        self.update_preview()
        QMessageBox.information(self, "Receipt Settings", "收据与打印机设置已保存。")

    def _current_qr_path(self) -> Path | None:
        paths = AppPaths.default()
        candidates: list[Path] = []
        if self._qr_source_path:
            source = Path(self._qr_source_path)
            candidates.append(source)
            candidates.append(paths.assets / source.name)
        candidates.extend(
            (
                paths.receipt_qr_image,
                paths.assets / "receipt_qr.jpg",
                paths.assets / RECEIPT_QR_IMAGE_NAME,
            )
        )
        for candidate in candidates:
            if candidate.is_file() and candidate.stat().st_size > 0:
                return candidate
        return None

    def _refresh_qr_preview(self) -> None:
        path = self._current_qr_path()
        if path is None:
            self.qr_preview.setPixmap(QPixmap())
            self.qr_preview.setText("No QR image")
            return
        pixmap = QPixmap(str(path))
        if pixmap.isNull():
            self.qr_preview.setPixmap(QPixmap())
            self.qr_preview.setText("Invalid image")
            return
        self.qr_preview.setText("")
        self.qr_preview.setPixmap(
            pixmap.scaled(
                self.qr_preview.size(),
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
        )

    def upload_qr(self) -> None:
        selected, _ = QFileDialog.getOpenFileName(
            self,
            "Select DuitNow Payment QR Image / 选择 DuitNow 收款码",
            "",
            "Images (*.png *.jpg *.jpeg);;PNG (*.png);;JPEG (*.jpg *.jpeg)",
            options=QFileDialog.Option.DontUseNativeDialog,
        )
        if not selected:
            return
        self._qr_source_path = selected
        self.qr_enabled.setChecked(True)
        self._refresh_qr_preview()
        self.update_preview()

    def clear_qr(self) -> None:
        paths = AppPaths.default()
        for candidate in (
            paths.receipt_qr_image,
            paths.assets / "receipt_qr.jpg",
            paths.assets / RECEIPT_QR_IMAGE_NAME,
        ):
            candidate.unlink(missing_ok=True)
        self._qr_source_path = None
        self.qr_enabled.setChecked(False)
        self._refresh_qr_preview()
        self.update_preview()

    def test_print(self) -> None:
        from PySide6.QtCore import QSizeF
        from PySide6.QtGui import QPageSize, QTextDocument
        from PySide6.QtPrintSupport import QPrinter, QPrinterInfo

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setPageSize(
            QPageSize(QSizeF(80, 297), QPageSize.Unit.Millimeter, "80mm")
        )
        test_pdf = os.environ.get("CNKH_POS_TEST_PRINT_PDF")
        output_path = None
        if test_pdf:
            paths = AppPaths.default()
            paths.ensure_directories()
            output_path = paths.exports / "receipt-settings-test.pdf"
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(str(output_path))
        else:
            available = set(QPrinterInfo.availablePrinterNames())
            selected = self.printer.currentData()
            settings = {
                "printer_mode": (
                    "DEFAULT"
                    if selected == WINDOWS_DEFAULT_PRINTER
                    else ("NAMED" if selected is not None else "UNCONFIGURED")
                ),
                "printer_name": (
                    ""
                    if selected in (None, WINDOWS_DEFAULT_PRINTER)
                    else str(selected)
                ),
            }
            try:
                target = resolve_printer_target(
                    settings,
                    available_printers=available,
                    default_printer_available=not QPrinterInfo.defaultPrinter().isNull(),
                )
            except RuntimeError as exc:
                QMessageBox.warning(self, "Test Print", str(exc))
                return
            if target is not None:
                printer.setPrinterName(target)
        document = QTextDocument()
        document.setPlainText(self.preview.toPlainText())
        document.print_(printer)
        if printer.printerState() == QPrinter.PrinterState.Error:
            QMessageBox.warning(
                self, "Test Print", "打印机在发送测试小票时报告错误。"
            )
            return
        if output_path is not None and (
            not output_path.is_file() or output_path.stat().st_size == 0
        ):
            QMessageBox.warning(self, "Test Print", "测试 PDF 未成功建立。")
            return
        QMessageBox.information(
            self,
            "Test Print",
            f"测试小票已输出：{output_path}"
            if output_path
            else f"测试小票已发送到：{self.printer.currentText()}",
        )


class DocumentPrefixesWidget(QWidget):
    def __init__(self, database: Database, user: AuthenticatedUser):
        super().__init__()
        self.database = database
        self.user = user
        layout = QVBoxLayout(self)
        layout.addWidget(
            QLabel(
                "单号前缀会用于之后新建立的收据、进货单、退货单和盘点单；旧单号不会改变。"
            )
        )
        form = QFormLayout()
        self.controls: dict[str, QLineEdit] = {}
        labels = {
            "RECEIPT": "Receipt / 收据",
            "PURCHASE": "Purchase / 进货",
            "RETURN": "Return / 退货",
            "STOCKTAKE": "Stocktake / 盘点",
        }
        conn = database.connect(readonly=True)
        try:
            current = document_prefixes(conn)
        finally:
            conn.close()
        for key, default in DEFAULT_DOCUMENT_PREFIXES.items():
            control = QLineEdit(current.get(key, default))
            control.setMaxLength(12)
            control.setPlaceholderText(default)
            self.controls[key] = control
            form.addRow(labels[key], control)
        layout.addLayout(form)
        save = QPushButton("保存单号前缀")
        save.setObjectName("SuccessButton")
        save.clicked.connect(self.save)
        self.save_button = save
        layout.addWidget(save, alignment=Qt.AlignmentFlag.AlignLeft)
        layout.addStretch(1)

    def save(self) -> None:
        try:
            with self.database.transaction() as conn:
                normalized = save_document_prefixes(
                    conn,
                    {key: control.text() for key, control in self.controls.items()},
                    admin_id=self.user.id,
                )
        except Exception as exc:
            QMessageBox.warning(self, "Document Prefixes", str(exc))
            return
        for key, value in normalized.items():
            self.controls[key].setText(value)
        QMessageBox.information(self, "Document Prefixes", "单号前缀已保存。")



class LanSyncSettingsWidget(QWidget):
    """Start/stop local LAN sync HTTP API for phone companions."""

    def __init__(self, database: Database, user: AuthenticatedUser, parent=None):
        super().__init__(parent)
        self.database = database
        self.user = user
        root = QVBoxLayout(self)
        root.addWidget(QLabel(
            "局域网同步 / LAN Sync — 手机与电脑同一 Wi‑Fi，无需云端。\n"
            "手机 Settings → Sync 填写下方地址与密钥。"
        ))
        form = QFormLayout()
        self.port = QLineEdit(str(DEFAULT_PORT))
        self.token = QLineEdit()
        self.token.setPlaceholderText("可选共享密钥 / optional shared secret")
        self.endpoint = QLabel("—")
        self.endpoint.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self.status = QLabel("Stopped")
        form.addRow("Port", self.port)
        form.addRow("Token / PIN", self.token)
        form.addRow("Endpoint", self.endpoint)
        form.addRow("Status", self.status)
        root.addLayout(form)
        row = QHBoxLayout()
        start = QPushButton("启动同步服务 / Start Sync Server")
        start.setObjectName("SuccessButton")
        start.clicked.connect(self.start_server)
        stop = QPushButton("停止 / Stop")
        stop.setObjectName("DangerButton")
        stop.clicked.connect(self.stop_server)
        row.addWidget(start)
        row.addWidget(stop)
        show_qr = QPushButton("显示配对二维码 / Show pair QR")
        show_qr.setObjectName("PrimaryButton")
        show_qr.clicked.connect(
            lambda: open_sync_pair_dialog(self, self.database)
        )
        row.addWidget(show_qr)
        root.addLayout(row)
        tip = QLabel(
            "防火墙请允许入站 TCP（默认 8787）。手机与 PC 须同一局域网。\n"
            "协议见仓库 LAN_SYNC.md。"
        )
        tip.setWordWrap(True)
        root.addWidget(tip)
        root.addStretch(1)
        self._refresh()

    def _refresh(self) -> None:
        srv = get_active_server()
        if srv is not None and srv.running:
            self.status.setText("Running")
            self.endpoint.setText(srv.endpoint)
            self.port.setText(str(srv.port))
        else:
            self.status.setText("Stopped")
            self.endpoint.setText(f"http://{detect_lan_ip()}:{self.port.text().strip() or DEFAULT_PORT}")

    def start_server(self) -> None:
        try:
            port = int(self.port.text().strip() or DEFAULT_PORT)
        except ValueError:
            QMessageBox.warning(self, "LAN Sync", "无效端口 / Invalid port")
            return
        try:
            srv = start_global_server(
                self.database, port=port, token=self.token.text().strip()
            )
            self._refresh()
            QMessageBox.information(
                self,
                "LAN Sync",
                f"已启动 / Started\n{srv.endpoint}\n"
                "请用手机顶栏「扫码配对」扫描二维码。",
            )
        except Exception as exc:
            QMessageBox.critical(self, "LAN Sync", str(exc))

    def stop_server(self) -> None:
        stop_global_server()
        self._refresh()
        QMessageBox.information(self, "LAN Sync", "已停止 / Stopped")


def _save_setting(database: Database, user: AuthenticatedUser, key: str, value: str) -> None:
    import json as _json

    from cnkh_pos.database.migrations import utc_now_text
    with database.transaction() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO settings(key,value_json,updated_at,updated_by) VALUES (?,?,?,?)",
            (key, _json.dumps(value), utc_now_text(), user.id),
        )
    from PySide6.QtWidgets import QMessageBox as _MB
    _MB.information(None, "Settings", f"已保存 {key}={value}")


class SettingsPage(QWidget):
    def __init__(self, database: Database, user: AuthenticatedUser):
        super().__init__()
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 20)
        header = QHBoxLayout()
        title = QLabel("系统设置")
        title.setObjectName("PageTitle")
        categories = QPushButton("Category Management / 分类管理")
        categories.setObjectName("PrimaryButton")
        categories.clicked.connect(lambda: CategoryDialog(database, user, self).exec())
        header.addWidget(title)
        header.addStretch(1)
        header.addWidget(categories)
        root.addLayout(header)
        tabs = QTabWidget()
        tabs.addTab(ReceiptSettingsWidget(database, user), "Receipt Settings")
        tabs.addTab(QuickAmountsWidget(database), "Quick Amount Settings")
        tabs.addTab(DocumentPrefixesWidget(database, user), "Document Prefixes")
        general = QWidget()
        general_layout = QVBoxLayout(general)
        general_layout.addWidget(QLabel("界面语言：中文 / English（固定双语）"))
        general_layout.addWidget(
            QLabel("Staff POS 开机自动启动可在安装器选项中启用或关闭。")
        )
        general_layout.addWidget(QLabel("数据库备份默认保留最近 30 份。"))
        general_layout.addWidget(QLabel("—"))
        general_layout.addWidget(QLabel("库存不足策略 / Stock gate（settings.key=stock_policy）"))
        stock_row = QHBoxLayout()
        self._stock_warn = QPushButton("警告可继续 warn")
        self._stock_block = QPushButton("阻止结账 block")
        self._stock_warn.clicked.connect(
            lambda: _save_setting(database, user, "stock_policy", "warn")
        )
        self._stock_block.clicked.connect(
            lambda: _save_setting(database, user, "stock_policy", "block")
        )
        stock_row.addWidget(self._stock_warn)
        stock_row.addWidget(self._stock_block)
        general_layout.addLayout(stock_row)
        general_layout.addWidget(QLabel("缺货推送阈值 / Low-stock threshold（low_stock_threshold，默认 10）"))
        self._low_stock_thr = QLineEdit("10")
        low_save = QPushButton("保存阈值")
        low_save.clicked.connect(
            lambda: _save_setting(
                database, user, "low_stock_threshold", self._low_stock_thr.text().strip() or "10"
            )
        )
        low_row = QHBoxLayout()
        low_row.addWidget(self._low_stock_thr)
        low_row.addWidget(low_save)
        general_layout.addLayout(low_row)
        general_layout.addWidget(QLabel("商品图片 / Product images（product_images_enabled）"))
        img_row = QHBoxLayout()
        img_on = QPushButton("开启 on")
        img_off = QPushButton("关闭 off")
        img_on.clicked.connect(lambda: _save_setting(database, user, "product_images_enabled", "1"))
        img_off.clicked.connect(lambda: _save_setting(database, user, "product_images_enabled", "0"))
        img_row.addWidget(img_on)
        img_row.addWidget(img_off)
        general_layout.addLayout(img_row)
        general_layout.addWidget(QLabel("挂单超时分钟 / Hold timeout（hold_timeout_minutes）"))
        self._hold_timeout = QLineEdit("30")
        hold_save = QPushButton("保存超时")
        hold_save.clicked.connect(
            lambda: _save_setting(
                database, user, "hold_timeout_minutes", self._hold_timeout.text().strip() or "30"
            )
        )
        hold_row = QHBoxLayout()
        hold_row.addWidget(self._hold_timeout)
        hold_row.addWidget(hold_save)
        general_layout.addLayout(hold_row)
        # load current
        try:
            import json as _json

            conn = database.connect(readonly=True)
            stock_row = conn.execute(
                "SELECT value_json FROM settings WHERE key='stock_policy'"
            ).fetchone()
            hold = conn.execute(
                "SELECT value_json FROM settings WHERE key='hold_timeout_minutes'"
            ).fetchone()
            low = conn.execute(
                "SELECT value_json FROM settings WHERE key='low_stock_threshold'"
            ).fetchone()
            conn.close()
            if low and low[0]:
                try:
                    self._low_stock_thr.setText(str(_json.loads(low[0])))
                except Exception:
                    self._low_stock_thr.setText(str(low[0]).strip('"') or "10")
            if stock_row and stock_row[0]:
                try:
                    policy = str(_json.loads(stock_row[0]))
                except Exception:
                    policy = str(stock_row[0]).strip('"') or "warn"
                # Visual cue only — buttons still save on click
                if policy == "block":
                    self._stock_block.setDefault(True)
                else:
                    self._stock_warn.setDefault(True)
            if hold and hold[0]:
                try:
                    self._hold_timeout.setText(str(_json.loads(hold[0])))
                except Exception:
                    self._hold_timeout.setText(str(hold[0]).strip('"') or "30")
        except Exception:
            pass
        general_layout.addStretch(1)
        tabs.addTab(general, "General")
        tabs.addTab(LanSyncSettingsWidget(database, user), "LAN Sync / 局域网同步")
        root.addWidget(tabs)


class DailyClosingPage(QWidget):
    def __init__(self, database: Database, user: AuthenticatedUser):
        super().__init__()
        self.database = database
        self.user = user
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 24)
        title = QLabel("Daily Cash Closing / 每日收银结算")
        title.setObjectName("PageTitle")
        root.addWidget(title)
        form = QFormLayout()
        self.opening = QDoubleSpinBox()
        self.opening.setDecimals(2)
        self.opening.setMaximum(99999999)
        self.system = QLineEdit("0.00")
        self.system.setReadOnly(True)
        self.actual = QDoubleSpinBox()
        self.actual.setDecimals(2)
        self.actual.setMaximum(99999999)
        self.variance = QLabel("RM 0.00")
        self.note = QTextEdit()
        self.opening.valueChanged.connect(self.load_system_cash)
        self.actual.valueChanged.connect(self.update_variance)
        form.addRow("Opening Cash / 开档现金", self.opening)
        form.addRow("Expected Cash / 系统应有", self.system)
        form.addRow("Actual Cash", self.actual)
        form.addRow("Variance", self.variance)
        form.addRow("Notes", self.note)
        root.addLayout(form)
        complete = QPushButton("完成日结")
        complete.setObjectName("CheckoutButton")
        complete.clicked.connect(self.complete)
        root.addWidget(complete)
        root.addStretch(1)
        self.load_system_cash()

    def load_system_cash(self) -> None:
        cents = DailyClosingService(self.database).system_cash(
            business_date=date.today(),
            opening_cash_cents=rm_to_cents(self.opening.value()),
        )
        self.system.setText(f"{cents / 100:.2f}")
        self.update_variance()

    def update_variance(self) -> None:
        try:
            system = rm_to_cents(self.system.text())
            actual = rm_to_cents(self.actual.value())
            self.variance.setText(format_myr(actual - system))
        except ValueError:
            self.variance.setText("—")

    def complete(self) -> None:
        try:
            DailyClosingService(self.database).complete(
                business_date=date.today(),
                cashier_id=self.user.id,
                opening_cash_cents=rm_to_cents(self.opening.value()),
                actual_cash_cents=rm_to_cents(self.actual.value()),
                note=self.note.toPlainText(),
            )
        except Exception as exc:
            QMessageBox.warning(self, "Daily Closing", str(exc))
            return
        QMessageBox.information(self, "Daily Closing", "日结已保存。")


class ReportsPage(QWidget):
    def __init__(self, database: Database):
        super().__init__()
        self.database = database
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 20, 22, 20)
        title = QLabel("Reports / Monthly Summary")
        title.setObjectName("PageTitle")
        root.addWidget(title)
        filters = QHBoxLayout()
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        today = QDate.currentDate()
        self.start_date.setDate(QDate(today.year(), today.month(), 1))
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(today)
        refresh = QPushButton("Refresh / 刷新")
        refresh.setObjectName("PrimaryButton")
        refresh.clicked.connect(self.refresh)
        filters.addWidget(QLabel("From"))
        filters.addWidget(self.start_date)
        filters.addWidget(QLabel("To"))
        filters.addWidget(self.end_date)
        filters.addWidget(refresh)
        filters.addStretch(1)
        root.addLayout(filters)
        self.summary = QLabel()
        self.summary.setStyleSheet("font-size:17px; line-height:1.5;")
        root.addWidget(self.summary)
        export = QPushButton("Export Excel")
        export.setObjectName("PrimaryButton")
        export.clicked.connect(self.export_excel)
        self.export_button = export
        root.addWidget(export, alignment=Qt.AlignmentFlag.AlignLeft)
        root.addStretch(1)
        self.refresh()

    def _date_range(self) -> tuple[str, str]:
        start = self.start_date.date().toString("yyyy-MM-dd")
        end = self.end_date.date().toString("yyyy-MM-dd")
        if start > end:
            raise ValueError("start date cannot be after end date")
        return start, end

    def refresh(self) -> bool:
        try:
            start, end = self._date_range()
        except ValueError as exc:
            QMessageBox.warning(self, "Reports", str(exc))
            return False
        try:
            result = ReportService(self.database).summary(
                start_date=start, end_date=end
            )
        except Exception as exc:
            QMessageBox.warning(self, "Reports", str(exc))
            return False
        self.summary.setText(
            f"Period: {start} to {end}\nSales: {format_myr(result.sales_cents)}\n"
            f"Gross Profit after Discounts: {format_myr(result.gross_profit_cents)}\n"
            f"Transaction Count: {result.transaction_count}\n"
            f"Purchases: {format_myr(result.purchases_cents)}\n"
            f"Current Customer Receivables: {format_myr(result.current_receivable_cents)}\n"
            f"Current Supplier Payables: {format_myr(result.current_payable_cents)}"
        )
        return True

    def export_excel(self) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        try:
            start, end = self._date_range()
        except ValueError as exc:
            QMessageBox.warning(self, "Reports", str(exc))
            return
        if not self.refresh():
            return
        paths = AppPaths.default()
        paths.ensure_directories()
        target = paths.exports / f"CNKH_POS_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(["CNKH Hardware POS V5", "Monthly Summary"])
        for line in self.summary.text().splitlines():
            label, _, value = line.partition(":")
            summary.append([label, value.strip()])
        summary["A1"].font = Font(bold=True, color="FFFFFF")
        summary["B1"].font = Font(bold=True, color="FFFFFF")
        summary["A1"].fill = PatternFill("solid", fgColor="0B2A53")
        summary["B1"].fill = PatternFill("solid", fgColor="0B2A53")
        summary.column_dimensions["A"].width = 28
        summary.column_dimensions["B"].width = 24

        conn = self.database.connect(readonly=True)
        try:
            datasets = (
                (
                    "Sales",
                    ["Receipt No", "Sold At", "Total (sen)", "Payment", "Customer"],
                    """SELECT s.receipt_no,s.sold_at,s.total_cents,s.payment_method,
                              COALESCE(c.name,'Walk-In Customer')
                       FROM sales s LEFT JOIN customers c ON c.id=s.customer_id
                       WHERE s.is_deleted=0 AND substr(s.sold_at,1,10) BETWEEN ? AND ?
                       ORDER BY s.sold_at DESC""",
                    (start, end),
                ),
                (
                    "Purchases",
                    ["Purchase No", "Purchased At", "Total (sen)", "Paid (sen)", "Status"],
                    """SELECT purchase_no,purchased_at,total_cents,paid_cents,status
                       FROM purchases WHERE is_deleted=0
                       AND substr(purchased_at,1,10) BETWEEN ? AND ?
                       ORDER BY purchased_at DESC""",
                    (start, end),
                ),
                (
                    "Returns",
                    [
                        "Return No",
                        "Returned At",
                        "Original Receipt",
                        "Refund (sen)",
                        "Refund Method",
                        "Reason",
                    ],
                    """SELECT r.return_no,r.returned_at,s.receipt_no,r.total_cents,
                              r.refund_method,r.reason
                       FROM sale_returns r JOIN sales s ON s.id=r.sale_id
                       WHERE substr(r.returned_at,1,10) BETWEEN ? AND ?
                       ORDER BY r.returned_at DESC""",
                    (start, end),
                ),
                (
                    "Customer Debts",
                    ["Customer", "Original (sen)", "Balance (sen)", "Status", "Opened At"],
                    """SELECT c.name,d.original_cents,d.balance_cents,d.status,d.opened_at
                       FROM customer_debts d JOIN customers c ON c.id=d.customer_id
                       WHERE substr(d.opened_at,1,10) BETWEEN ? AND ?
                       ORDER BY d.opened_at DESC""",
                    (start, end),
                ),
            )
            for name, headers, query, params in datasets:
                sheet = workbook.create_sheet(name)
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1769E0")
                for row in conn.execute(query, params):
                    sheet.append(list(row))
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions
                for column in sheet.columns:
                    width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
                    sheet.column_dimensions[column[0].column_letter].width = width
        finally:
            conn.close()
        workbook.save(target)
        QMessageBox.information(self, "Export Excel", f"报表已导出：{target}")
