from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from cnkh_pos.database.connection import Database
from cnkh_pos.database.migrations import utc_now_text
from cnkh_pos.services.catalog import CatalogService, ProductInput, is_valid_ean13
from cnkh_pos.services.money import rm_to_cents


HEADERS = [
    "Product Name",
    "Aliases",
    "Category",
    "SKU",
    "Cost RM",
    "Selling Price RM",
    "Stock",
    "Unit",
    "Location",
    "Low Stock Level",
    "Barcode",
]


@dataclass(slots=True)
class ImportRow:
    row_number: int
    values: dict[str, object]
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class ImportSummary:
    success: int
    skipped: int
    errors: int


class ExcelImportService:
    def __init__(self, database: Database):
        self.database = database

    @staticmethod
    def create_template(path: Path) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(HEADERS)
        sheet.append(
            [
                "PVC Pipe 20mm",
                "paip pvc; 管子",
                "Water Pipe",
                "PIPE20",
                2.0,
                4.5,
                80,
                "meter",
                "Rack B2",
                10,
                "",
            ]
        )
        fill = PatternFill("solid", fgColor="1769E0")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:K2"
        for column, width in zip(
            "ABCDEFGHIJK", [28, 24, 20, 16, 14, 18, 12, 14, 18, 18, 20], strict=True
        ):
            sheet.column_dimensions[column].width = width
        workbook.save(path)
        return path

    def preview(self, path: Path) -> list[ImportRow]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        missing_headers = [name for name in HEADERS if name not in headers]
        if missing_headers:
            raise ValueError("missing template columns: " + ", ".join(missing_headers))
        seen_barcode: set[str] = set()
        seen_sku: set[str] = set()
        seen_name: set[str] = set()
        result: list[ImportRow] = []
        conn = self.database.connect(readonly=True)
        try:
            for row_number, cells in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                values = dict(zip(headers, cells, strict=False))
                if not any(value not in (None, "") for value in values.values()):
                    continue
                item = ImportRow(row_number, values)
                name = str(values.get("Product Name") or "").strip()
                sku = str(values.get("SKU") or "").strip()
                barcode = str(values.get("Barcode") or "").strip().replace(".0", "")
                if not name:
                    item.errors.append("Missing Required Field: Product Name")
                name_key = name.casefold()
                if name_key and (
                    name_key in seen_name
                    or conn.execute(
                        "SELECT 1 FROM products WHERE name=? COLLATE NOCASE AND is_deleted=0",
                        (name,),
                    ).fetchone()
                ):
                    item.errors.append("Duplicate Name")
                if sku and (
                    sku.casefold() in seen_sku
                    or conn.execute(
                        "SELECT 1 FROM products WHERE sku=? COLLATE NOCASE", (sku,)
                    ).fetchone()
                ):
                    item.errors.append("Duplicate SKU")
                if barcode:
                    if not is_valid_ean13(barcode):
                        item.errors.append("Invalid Barcode")
                    if (
                        barcode in seen_barcode
                        or conn.execute(
                            "SELECT 1 FROM products WHERE barcode=?", (barcode,)
                        ).fetchone()
                    ):
                        item.errors.append("Duplicate Barcode")
                seen_name.add(name_key)
                if sku:
                    seen_sku.add(sku.casefold())
                if barcode:
                    seen_barcode.add(barcode)
                for money_column in ("Cost RM", "Selling Price RM"):
                    try:
                        rm_to_cents(str(values.get(money_column) or "0"))
                    except ValueError:
                        item.errors.append(f"Invalid {money_column}")
                result.append(item)
        finally:
            conn.close()
            workbook.close()
        return result

    def commit(self, rows: list[ImportRow], *, admin_id: int) -> ImportSummary:
        service = CatalogService(self.database)
        success = skipped = errors = 0
        category_cache: dict[str, int] = {}
        for row in rows:
            if row.errors:
                errors += 1
                continue
            values = row.values
            try:
                category_name = str(values.get("Category") or "").strip()
                category_id = None
                if category_name:
                    if category_name not in category_cache:
                        with self.database.transaction() as conn:
                            existing = conn.execute(
                                "SELECT id FROM categories WHERE name=? COLLATE NOCASE",
                                (category_name,),
                            ).fetchone()
                            if existing is None:
                                now = utc_now_text()
                                category_cache[category_name] = int(
                                    conn.execute(
                                        "INSERT INTO categories(name, created_at, updated_at) VALUES (?, ?, ?)",
                                        (category_name, now, now),
                                    ).lastrowid
                                )
                            else:
                                category_cache[category_name] = int(existing["id"])
                    category_id = category_cache[category_name]
                service.add_product(
                    ProductInput(
                        name=str(values.get("Product Name") or ""),
                        aliases=str(values.get("Aliases") or ""),
                        category_id=category_id,
                        sku=str(values.get("SKU") or "") or None,
                        cost_cents=rm_to_cents(str(values.get("Cost RM") or "0")),
                        selling_price_cents=rm_to_cents(
                            str(values.get("Selling Price RM") or "0")
                        ),
                        stock=str(values.get("Stock") or "0"),
                        unit=str(values.get("Unit") or "pcs"),
                        location=str(values.get("Location") or ""),
                        low_stock=str(values.get("Low Stock Level") or "0"),
                        barcode=str(values.get("Barcode") or "").replace(".0", "")
                        or None,
                    ),
                    admin_id=admin_id,
                )
                success += 1
            except Exception as exc:
                row.errors.append(str(exc))
                skipped += 1
        return ImportSummary(success, skipped, errors)
