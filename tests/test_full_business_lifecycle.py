from __future__ import annotations

import json
import os
import re
import sys
import tempfile
import traceback
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication, QPushButton, QTabWidget

from cnkh_pos.database.bootstrap import bootstrap_database
from cnkh_pos.database.connection import Database
from cnkh_pos.services.auth import AuthService
from cnkh_pos.services.backup import BackupService
from cnkh_pos.services.barcode_labels import (
    get_label_profile,
    load_product_label,
    render_product_label_pdf,
)
from cnkh_pos.services.catalog import (
    CatalogService,
    CategoryService,
    ProductInput,
    is_valid_ean13,
)
from cnkh_pos.services.checkout_rounding import (
    RoundedReturnService,
    RoundedSalesService,
)
from cnkh_pos.services.daily_closing import DailyClosingService
from cnkh_pos.services.discounts import discount_cents_from_value
from cnkh_pos.services.entities import EntityInput, EntityService
from cnkh_pos.services.excel_import import ExcelImportService
from cnkh_pos.services.held_orders import HeldOrderService, cart_state_from_held_payload
from cnkh_pos.services.payments import CustomerPaymentService, SupplierPaymentService
from cnkh_pos.services.printing import PrintingService
from cnkh_pos.services.purchases import PurchaseLine, PurchaseService
from cnkh_pos.services.reports import ReportService
from cnkh_pos.services.restore import RestoreService
from cnkh_pos.services.sales import SaleLine, SalesService
from cnkh_pos.services.stocktake import StocktakeService
from cnkh_pos.ui.admin import AdminWindow
from cnkh_pos.ui.admin.enhanced_data_pages import ProductDialogWithBarcodeMode
from cnkh_pos.ui.dialogs.rounded_checkout import RoundedCheckoutDialog
from cnkh_pos.ui.staff import StaffWindow
from cnkh_pos.ui.theme import apply_theme

pytestmark = pytest.mark.skipif(
    sys.platform != "win32",
    reason="final business lifecycle acceptance is intentionally Windows-only",
)

REPORT_PATH = Path("self-test-artifacts") / "full-business-acceptance.json"


def _write_report(status: str, checks: list[dict[str, object]], error: str = "") -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(
        json.dumps(
            {"status": status, "checks": checks, "error": error},
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )


def _stock(database: Database, product_id: int) -> Decimal:
    conn = database.connect(readonly=True)
    try:
        row = conn.execute(
            "SELECT stock_decimal FROM products WHERE id=?", (product_id,)
        ).fetchone()
        if row is None:
            raise LookupError("product not found")
        return Decimal(str(row[0]))
    finally:
        conn.close()


def test_full_windows_business_lifecycle() -> None:
    checks: list[dict[str, object]] = []

    def passed(name: str, detail: object = "PASS") -> None:
        checks.append({"name": name, "status": "PASS", "detail": str(detail)})

    try:
        app = QApplication.instance() or QApplication([])
        apply_theme(app)
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            database = Database(root / "hardware_pos.db")
            backups = root / "backups"
            bootstrap_database(database.path, backups)

            with database.transaction() as conn:
                admin_id = AuthService.create_user(
                    conn,
                    username="business-admin",
                    display_name="Business Admin",
                    password="AdminPass123!",
                    role="ADMIN",
                    permissions={},
                    admin_id=None,
                )
                staff_id = AuthService.create_user(
                    conn,
                    username="business-staff",
                    display_name="Business Staff",
                    password="StaffPass123!",
                    role="STAFF",
                    permissions={
                        "apply_discount": True,
                        "manage_quick_amounts": True,
                        "reprint_receipt": True,
                    },
                    admin_id=admin_id,
                )
                admin_user = AuthService.authenticate(
                    conn,
                    "business-admin",
                    "AdminPass123!",
                    required_role="ADMIN",
                )
                staff_user = AuthService.authenticate(
                    conn,
                    "business-staff",
                    "StaffPass123!",
                    required_role="STAFF",
                )
            assert admin_user.id == admin_id and staff_user.id == staff_id
            passed("users.create_and_authenticate")

            categories = CategoryService(database)
            category_id = categories.add("Hardware", admin_id=admin_id)
            categories.rename(category_id, "Hardware Main", admin_id=admin_id)
            passed("categories.add_and_rename")

            customers = EntityService(database, "customers")
            customer_id = customers.add(
                EntityInput("Walk-in Credit Customer", phone="0123456789"),
                admin_id=admin_id,
            )
            customers.update(
                customer_id,
                EntityInput("Account Customer", phone="0123456789", notes="verified"),
                admin_id=admin_id,
            )
            suppliers = EntityService(database, "suppliers")
            supplier_id = suppliers.add(
                EntityInput(
                    "Main Supplier",
                    phone="0312345678",
                    email="supplier@example.invalid",
                ),
                admin_id=admin_id,
            )
            suppliers.update(
                supplier_id,
                EntityInput(
                    "Main Supplier Updated",
                    phone="0312345678",
                    email="supplier@example.invalid",
                    notes="preferred",
                ),
                admin_id=admin_id,
            )
            delete_supplier_id = suppliers.add(
                EntityInput("Disposable Supplier"), admin_id=admin_id
            )
            passed("customers_and_suppliers.add_update")

            catalog = CatalogService(database)
            auto_product = catalog.add_product(
                ProductInput(
                    name="Auto Barcode Hammer",
                    category_id=category_id,
                    sku="HAM-AUTO",
                    cost_cents=300,
                    selling_price_cents=1067,
                    stock="10",
                    unit="pcs",
                    location="A1",
                    low_stock="2",
                ),
                admin_id=admin_id,
            )
            manual_product = catalog.add_product(
                ProductInput(
                    name="Manual Barcode Screw",
                    category_id=category_id,
                    sku="SCR-MANUAL",
                    barcode="CNKH-MANUAL-001",
                    cost_cents=250,
                    selling_price_cents=500,
                    stock="5",
                    unit="box",
                    location="B2",
                    low_stock="1",
                ),
                admin_id=admin_id,
            )
            disposable_product = catalog.add_product(
                ProductInput(name="Disposable Product", selling_price_cents=200, stock="0"),
                admin_id=admin_id,
            )
            conn = database.connect(readonly=True)
            try:
                auto_barcode = str(
                    conn.execute(
                        "SELECT barcode FROM products WHERE id=?", (auto_product,)
                    ).fetchone()[0]
                )
                manual_barcode = str(
                    conn.execute(
                        "SELECT barcode FROM products WHERE id=?", (manual_product,)
                    ).fetchone()[0]
                )
            finally:
                conn.close()
            assert is_valid_ean13(auto_barcode)
            assert manual_barcode == "CNKH-MANUAL-001"
            passed("products.auto_and_manual_barcode", auto_barcode)

            catalog.update_product(
                auto_product,
                ProductInput(
                    name="Auto Barcode Hammer Updated",
                    category_id=category_id,
                    sku="HAM-AUTO",
                    cost_cents=300,
                    selling_price_cents=1067,
                    stock="10",
                    unit="pcs",
                    location="A2",
                    low_stock="2",
                ),
                admin_id=admin_id,
            )
            passed("products.edit")

            suppliers.set_supplier_products(
                supplier_id, {auto_product, manual_product}, admin_id=admin_id
            )
            suppliers.set_supplier_products(
                delete_supplier_id, {disposable_product}, admin_id=admin_id
            )
            assert suppliers.supplier_product_ids(supplier_id) == {
                auto_product,
                manual_product,
            }
            passed("supplier_product_mapping")

            purchase = PurchaseService(database).create_purchase(
                supplier_id=supplier_id,
                lines=[PurchaseLine(auto_product, Decimal("5"), 400)],
                paid_cents=500,
                payment_method="CARD",
                operator_id=admin_id,
            )
            assert purchase.total_cents == 2000 and purchase.status == "PARTIAL"
            SupplierPaymentService(database).record_payment(
                purchase_id=purchase.purchase_id,
                amount_cents=1500,
                payment_method="CARD",
                note="settle purchase",
                operator_id=admin_id,
            )
            assert _stock(database, auto_product) == Decimal("15")
            conn = database.connect(readonly=True)
            try:
                status = conn.execute(
                    "SELECT status FROM purchases WHERE id=?", (purchase.purchase_id,)
                ).fetchone()[0]
            finally:
                conn.close()
            assert status == "PAID"
            passed("purchase.create_receive_and_pay")

            disposable_purchase = PurchaseService(database).create_purchase(
                supplier_id=delete_supplier_id,
                lines=[PurchaseLine(disposable_product, Decimal("2"), 100)],
                paid_cents=0,
                payment_method="CARD",
                operator_id=admin_id,
            )
            assert _stock(database, disposable_product) == Decimal("2")
            PurchaseService(database).delete_purchase(
                purchase_id=disposable_purchase.purchase_id, admin_id=admin_id
            )
            assert _stock(database, disposable_product) == Decimal("0")
            passed("purchase.delete_and_reverse_stock")

            held_service = HeldOrderService(database)
            held = held_service.hold(
                {
                    "items": [
                        {
                            "product_id": auto_product,
                            "quantity": "1",
                            "discount_cents": 10,
                        }
                    ]
                },
                cashier_id=staff_id,
            )
            retrieved = held_service.retrieve(held.id, cashier_id=staff_id)
            quantities, discounts = cart_state_from_held_payload(retrieved.payload)
            assert quantities[auto_product] == Decimal("1")
            assert discounts[auto_product] == 10
            passed("pos.hold_and_retrieve")

            staff_window = StaffWindow(database, staff_user)
            staff_window.cart_quantities = {
                auto_product: Decimal("1"),
                manual_product: Decimal("2"),
            }
            item_discount = discount_cents_from_value(
                1000, mode="FIXED", value="1.00"
            )
            staff_window.cart_discounts = {manual_product: item_discount}
            staff_window._rebuild_cart()
            raw_cart_total = staff_window._cart_total()
            assert raw_cart_total == 1967

            checkout = RoundedCheckoutDialog(raw_cart_total, parent=staff_window)
            checkout.checkout_discount_mode.setCurrentIndex(
                checkout.checkout_discount_mode.findData("PERCENT")
            )
            checkout.checkout_discount_value.setValue(7)
            assert checkout.discount_cents == 138
            assert checkout.discounted_total_cents == 1829
            assert checkout.total_cents == 1830
            lines = staff_window._sale_lines_with_order_discount(checkout.discount_cents)
            checkout.close()

            sale = RoundedSalesService(database).create_sale(
                lines=lines,
                payment_method="CASH",
                paid_cents=2000,
                cashier_id=staff_id,
            )
            assert sale.total_cents == 1830 and sale.change_cents == 170
            conn = database.connect(readonly=True)
            try:
                row = conn.execute(
                    "SELECT subtotal_cents,discount_cents,total_cents FROM sales WHERE id=?",
                    (sale.sale_id,),
                ).fetchone()
            finally:
                conn.close()
            assert tuple(row) == (2067, 238, 1830)
            passed("pos.checkout_percent_fixed_discount_and_rounding")

            printing = PrintingService(database)
            receipt = printing.receipt(sale.sale_id)
            receipt_text = printing.render_text(receipt)
            assert "TOTAL" in receipt_text and "RM 18.30" in receipt_text
            assert "DISCOUNT" in receipt_text and "RM 2.38" in receipt_text
            receipt_pdf = printing.render_pdf(receipt, root / "receipt.pdf")
            assert receipt_pdf.is_file() and receipt_pdf.stat().st_size > 500
            passed("receipt.text_and_80mm_pdf")

            stock_before_delete_sale = _stock(database, auto_product)
            delete_sale = RoundedSalesService(database).create_sale(
                lines=[SaleLine(auto_product, Decimal("2"), Decimal("2"))],
                payment_method="CASH",
                paid_cents=2200,
                cashier_id=staff_id,
            )
            conn = database.connect(readonly=True)
            try:
                delete_sale_item = int(
                    conn.execute(
                        "SELECT id FROM sale_items WHERE sale_id=?", (delete_sale.sale_id,)
                    ).fetchone()[0]
                )
            finally:
                conn.close()
            RoundedReturnService(database).create_return(
                sale_id=delete_sale.sale_id,
                quantities_by_sale_item={delete_sale_item: Decimal("1")},
                reason="business acceptance partial return",
                operator_id=admin_id,
                refund_method="CASH",
            )
            SalesService(database).delete_sale(
                sale_id=delete_sale.sale_id, admin_id=admin_id
            )
            assert _stock(database, auto_product) == stock_before_delete_sale
            conn = database.connect(readonly=True)
            try:
                assert (
                    conn.execute(
                        "SELECT COUNT(*) FROM sales WHERE id=?", (delete_sale.sale_id,)
                    ).fetchone()[0]
                    == 0
                )
            finally:
                conn.close()
            passed("sales.return_then_delete_with_stock_reversal")

            credit_sale = RoundedSalesService(database).create_sale(
                lines=[SaleLine(manual_product, Decimal("1"), Decimal("1"))],
                payment_method="CREDIT",
                paid_cents=0,
                cashier_id=staff_id,
                customer_id=customer_id,
            )
            assert credit_sale.total_cents == 500
            conn = database.connect(readonly=True)
            try:
                debt = conn.execute(
                    "SELECT id,balance_cents,status FROM customer_debts WHERE sale_id=?",
                    (credit_sale.sale_id,),
                ).fetchone()
            finally:
                conn.close()
            assert int(debt["balance_cents"]) == 500 and debt["status"] == "OPEN"
            CustomerPaymentService(database).record_payment(
                debt_id=int(debt["id"]),
                amount_cents=500,
                payment_method="CARD",
                note="settle customer",
                operator_id=admin_id,
            )
            conn = database.connect(readonly=True)
            try:
                debt_after = conn.execute(
                    "SELECT balance_cents,status FROM customer_debts WHERE id=?",
                    (debt["id"],),
                ).fetchone()
            finally:
                conn.close()
            assert tuple(debt_after) == (0, "CLOSED")
            passed("customer.credit_sale_and_payment")

            stocktake = StocktakeService(database)
            stocktake_id, stocktake_no = stocktake.create_draft(
                operator_id=admin_id, notes="business acceptance"
            )
            expected_physical = _stock(database, auto_product) + Decimal("1")
            stocktake.set_physical_count(
                stocktake_id=stocktake_id,
                product_id=auto_product,
                count=expected_physical,
            )
            stocktake.complete(stocktake_id=stocktake_id, operator_id=admin_id)
            assert _stock(database, auto_product) == expected_physical
            passed("stocktake.complete_with_variance", stocktake_no)

            today = date.today().isoformat()
            report = ReportService(database).summary(start_date=today, end_date=today)
            assert report.transaction_count == 2
            assert report.sales_cents == 2330
            assert report.current_receivable_cents == 0
            passed("reports.sales_profit_receivable")

            closing_service = DailyClosingService(database)
            system_cash = closing_service.system_cash(business_date=date.today())
            assert system_cash == 1830
            closing = closing_service.complete(
                business_date=date.today(),
                cashier_id=admin_id,
                actual_cash_cents=system_cash,
                note="business acceptance",
            )
            assert closing.variance_cents == 0
            passed("daily_cash_closing")

            product_label = load_product_label(database, auto_product)
            labels_pdf = render_product_label_pdf(
                product_label,
                get_label_profile("50x30"),
                4,
                root / "product-labels.pdf",
            )
            page_count = len(re.findall(rb"/Type\s*/Page\b", labels_pdf.read_bytes()))
            assert page_count == 4
            passed("barcode_labels.50x30_four_copies")

            template = ExcelImportService.create_template(root / "products.xlsx")
            workbook = load_workbook(template)
            sheet = workbook.active
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)
            sheet.append(
                [
                    "Excel Imported Product",
                    "excel alias",
                    "Imported Category",
                    "XLS-001",
                    1.25,
                    2.75,
                    2,
                    "pcs",
                    "X1",
                    1,
                    "",
                ]
            )
            workbook.save(template)
            workbook.close()
            excel_service = ExcelImportService(database)
            preview = excel_service.preview(template)
            assert len(preview) == 1 and not preview[0].errors
            import_summary = excel_service.commit(preview, admin_id=admin_id)
            assert import_summary.success == 1
            passed("excel.template_preview_import")

            dialog = ProductDialogWithBarcodeMode(database)
            assert dialog.barcode_mode is not None
            assert dialog.barcode_mode.currentData() == "AUTO"
            dialog.barcode_mode.setCurrentIndex(dialog.barcode_mode.findData("MANUAL"))
            assert dialog.barcode.isEnabled()
            dialog.close()
            passed("product_dialog.explicit_auto_manual_barcode_mode")

            categories.delete(category_id, admin_id=admin_id)
            customers.delete(customer_id, admin_id=admin_id)
            suppliers.delete(supplier_id, admin_id=admin_id)
            suppliers.delete(delete_supplier_id, admin_id=admin_id)
            deleted_count = catalog.delete_products(
                [manual_product, disposable_product], admin_id=admin_id
            )
            assert deleted_count == 2
            conn = database.connect(readonly=True)
            try:
                snapshot_count = int(
                    conn.execute(
                        "SELECT COUNT(*) FROM sale_items WHERE product_name_snapshot='Manual Barcode Screw'"
                    ).fetchone()[0]
                )
            finally:
                conn.close()
            assert snapshot_count >= 2
            passed("delete.customer_supplier_product_preserves_history")

            backup = BackupService(backups).create(
                database.path, reason="business_acceptance"
            ).path
            temporary_product = catalog.add_product(
                ProductInput(name="Temporary After Backup"), admin_id=admin_id
            )
            assert temporary_product > 0
            RestoreService(database, backups).restore(
                backup, admin_id=admin_id, password="AdminPass123!"
            )
            conn = database.connect(readonly=True)
            try:
                temporary_count = int(
                    conn.execute(
                        "SELECT COUNT(*) FROM products WHERE name='Temporary After Backup'"
                    ).fetchone()[0]
                )
            finally:
                conn.close()
            assert temporary_count == 0
            passed("backup_and_restore")

            ok, details = database.integrity_check()
            assert ok, details
            passed("database.integrity_check")

            admin_window = AdminWindow(database, admin_user)
            admin_window.show()
            app.processEvents()
            assert admin_window.isVisible()
            for key, index in admin_window.page_keys.items():
                admin_window.pages.setCurrentIndex(index)
                app.processEvents()
                assert admin_window.pages.currentIndex() == index, key
            catalog_tabs = admin_window.pages.widget(admin_window.page_keys["products"])
            assert isinstance(catalog_tabs, QTabWidget)
            assert catalog_tabs.count() == 3
            assert "Barcode Labels" in catalog_tabs.tabText(2)
            sales_page = admin_window.pages.widget(admin_window.page_keys["sales"])
            assert "删除销售记录" in [
                button.text() for button in sales_page.findChildren(QPushButton)
            ]
            admin_window.close()
            app.processEvents()
            passed("admin_gui.all_pages_and_new_controls")

            final_staff = StaffWindow(database, staff_user)
            final_staff.show()
            app.processEvents()
            assert final_staff.isVisible()
            assert final_staff.discount_button.isEnabled()
            assert final_staff.checkout_button.isEnabled()
            final_staff.close()
            app.processEvents()
            passed("staff_gui.opens_with_discount_and_checkout")

            conn = database.connect(readonly=True)
            try:
                audit_count = int(conn.execute("SELECT COUNT(*) FROM audit_logs").fetchone()[0])
            finally:
                conn.close()
            assert audit_count >= 20
            passed("audit_trail", audit_count)

        _write_report("PASS", checks)
    except BaseException as exc:
        checks.append(
            {
                "name": "exception",
                "status": "FAIL",
                "detail": f"{type(exc).__name__}: {exc}",
            }
        )
        _write_report(
            "FAIL",
            checks,
            "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)),
        )
        raise
